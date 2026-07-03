"""Merge every curated/discovery source into one deduplicated, prioritized candidate list.

Dedup key is the normalized product/service name. Occurrence count across
independent sources is used as a rough prominence/relevance signal for
priority ordering. Products already profiled in the original competitor
analysis template are flagged so they aren't re-researched.
"""

import json
import logging
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

CURATED_DIR = Path("data/curated")
JSTARTUP_DIR = Path("data/raw/jstartup")
ORIGINAL_TEMPLATE = Path("data/report/競合分析.xlsx")
OUTPUT_PATH = Path("data/master_candidates.json")

JSTARTUP_KEYWORDS = ["建設", "建築", "施工", "土木", "不動産", "ゼネコン", "archi"]


def normalize_name(name: str) -> str:
    """Normalize a product name into a dedup key (NFKC, trimmed, casefolded)."""
    return unicodedata.normalize("NFKC", name).strip().casefold()


def load_curated() -> list[dict[str, Any]]:
    """Load every data/curated/*.jsonl record as (name, category, source, url)."""
    records: list[dict[str, Any]] = []
    for path in sorted(CURATED_DIR.glob("*.jsonl")):
        with path.open(encoding="utf-8") as f:
            for line in f:
                row = json.loads(line)
                name = row.get("service_name")
                if not name:
                    continue
                category = (
                    row.get("category") or row.get("major_category") or row.get("sub_category")
                )
                records.append(
                    {
                        "name": name,
                        "category": category,
                        "source": row.get("source", path.stem),
                        "url": row.get("url"),
                    }
                )
    return records


def load_jstartup_filtered() -> list[dict[str, Any]]:
    """Load the latest J-Startup dump, keeping only construction-keyword hits."""
    paths = sorted(JSTARTUP_DIR.glob("*.jsonl"))
    if not paths:
        return []
    latest = paths[-1]
    records: list[dict[str, Any]] = []
    with latest.open(encoding="utf-8") as f:
        for line in f:
            row = json.loads(line)
            haystack = (row["name"] + " ".join(row.get("categories", []))).lower()
            if any(kw.lower() in haystack for kw in JSTARTUP_KEYWORDS):
                records.append(
                    {
                        "name": row["name"],
                        "category": "/".join(row.get("categories", [])),
                        "source": "jstartup",
                        "url": row.get("url"),
                    }
                )
    return records


def load_already_researched() -> set[str]:
    """Read product names already profiled in the original competitor analysis file."""
    if not ORIGINAL_TEMPLATE.exists():
        return set()
    wb = openpyxl.load_workbook(ORIGINAL_TEMPLATE, data_only=True)
    names: set[str] = set()
    ws = wb["競合リスト"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1, values_only=True):
        if row[0]:
            names.add(normalize_name(str(row[0])))
    ws = wb["AI活用動向"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=1, values_only=True):
        if row[0] and not str(row[0]).startswith("▼"):
            names.add(normalize_name(str(row[0])))
    return names


def is_already_researched(key: str, researched: set[str]) -> bool:
    """A candidate counts as covered if its name overlaps an already-researched one."""
    return any(key in r or r in key for r in researched)


def build_master_list() -> list[dict[str, Any]]:
    """Dedup all sources by normalized name and rank by source-occurrence count."""
    all_records = load_curated() + load_jstartup_filtered()
    researched = load_already_researched()

    grouped: dict[str, dict[str, Any]] = {}
    for rec in all_records:
        key = normalize_name(rec["name"])
        if key not in grouped:
            grouped[key] = {
                "name": rec["name"],
                "categories": set(),
                "sources": set(),
                "urls": set(),
            }
        entry = grouped[key]
        if rec["category"]:
            entry["categories"].add(rec["category"])
        entry["sources"].add(rec["source"])
        if rec["url"]:
            entry["urls"].add(rec["url"])

    candidates = []
    for key, entry in grouped.items():
        candidates.append(
            {
                "name": entry["name"],
                "categories": sorted(entry["categories"]),
                "sources": sorted(entry["sources"]),
                "occurrence_count": len(entry["sources"]),
                "urls": sorted(entry["urls"]),
                "already_researched": is_already_researched(key, researched),
            }
        )

    candidates.sort(key=lambda c: (c["already_researched"], -c["occurrence_count"], c["name"]))
    return candidates


def main() -> None:
    """Build the master candidate list and write it to OUTPUT_PATH."""
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    candidates = build_master_list()
    OUTPUT_PATH.write_text(json.dumps(candidates, ensure_ascii=False, indent=2), encoding="utf-8")
    new_count = sum(1 for c in candidates if not c["already_researched"])
    logger.info(
        "wrote %d candidates (%d new, %d already researched) to %s",
        len(candidates),
        new_count,
        len(candidates) - new_count,
        OUTPUT_PATH,
    )


if __name__ == "__main__":
    main()
